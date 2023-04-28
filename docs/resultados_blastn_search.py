# %%
from itertools import chain
from typing import Iterator
import os
from pathlib import Path
from time import perf_counter

# %%
here = os.getcwd()

# %%


def is_out(file: str) -> bool:
    extension = os.path.splitext(file)[1]
    if extension == '.out':
        return True
    else:
        return False


# %%
def get_out_folders(cwd:str) -> list[tuple[str, list[str]]]:
    """Busca en el directorio y los subdirectorios """
    dirs: Iterator[tuple[str, list, list]] = os.walk(cwd)
    out_folders = []
    for folder in dirs:
        folder_path, sub_dirs, files = folder
        out_files = list(filter(is_out, files))
        if out_files:
            out_folders.append((folder_path, out_files))
    return out_folders

# %%


def read_out_file(file_path: str) -> list[list[str]]:
    """Extrae los resultados de un .out"""
    lines: list[list[str]] = []
    with open(file_path, 'r') as out:
        for line in out.readlines():
            lines.append(line.split())
    return lines

# %%


def read_out_files(cwd: str) -> list[tuple[str, tuple[str, list[list[str]]]]]:
    '''Lee la carpeta que script y toda subcarpeta en busca de archivos .out

    Solo se retornan las carpetas y archivos que no estén vacíos
    here es desde se va a empezar a mirar'''
    print("Trabajando desde...", cwd)
    folders = get_out_folders(cwd)
    out_lines: list[tuple] = []
    for folder in folders:
        path, files = Path(folder[0]), folder[1]
        db = path.parts[-1].upper()

        folder_results = []
        for file in files:
            file_lines: list[list[str]] = []
            for line in read_out_file(os.path.join(path, file)):
                file_lines.append(line)
            if file_lines:
                file_results: tuple[str, list] = (file, file_lines)
                folder_results.append(file_results)

        if folder_results:
            out_lines.append((db, folder_results))
    return out_lines


# %%
# def write2xlsx(results:list[tuple[str,tuple[str, list[list[str]]]]], name: str):


def write2xlsx(results, name: str):
    """Escribe los resultados a un xlsx con el nombre dado"""
    from openpyxl.styles import Font
    if not results:
        print('sin resultados')
        return
    import openpyxl as xl
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Resultados"
    y: int = 1
    for folder in results:
        cell = 'A' + str(y)
        folder_name = folder[0]
        write_header(folder_name, ws, y)
        files = folder[1]
        for file in files:
            y += 2
            cell: str = 'A' + str(y)
            file_name = file[0]

            ws[cell] = file_name
            ws[cell].font = Font(name='Calibri', bold=True)
            lines = file[1]
            for line in lines:
                y += 1
                x = ord('A')
                for item in line:
                    cell = chr(x) + str(y)
                    x += 1
                    ws[cell] = item
        y += 5
    wb.save(name+'.xlsx')


def write_header(folder_name, worksheet, y):
    """Pone los headers bonitos como Julian"""
    from openpyxl.styles import Font
    headers = [folder_name, "", "Per. Ident",	"Longitud",	"Mismatch",
               "Gap Open",	"Q Start",	"Q end",	"Start",	"S end",	"E-Value",	"Bitscore"]
    x = ord('A')
    for item in headers:
        cell = chr(x) + str(y)
        worksheet[cell] = item
        worksheet[cell].font = Font(name='Calibri', bold=True)
        x += 1


# %%
def main():
    here = os.getcwd()
    i = perf_counter()
    dirs = read_out_files(here)
    f = perf_counter()

    print("Tiempo en leer los archivos: ", f-i)
    i = perf_counter()
    write2xlsx(dirs, "Resultados_blastn")
    f = perf_counter()
    print("Tiempo en escribir los archivos: ", f-i)


if "__main__" == __name__:
    main()
